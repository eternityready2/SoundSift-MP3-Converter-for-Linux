import os.path
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font


class Excel:
   lo_Inst_lst = []
   # Class constructor.
   # The scope of this calss is to have a data buffer, designed ass generic as possible to support multiple data elements.
   # Each instance of the class reprezents a line from the excel sheet (imported data or data to be written)
   # The mapping of the elements to the excel sheets is done as follows: el1 is the first column, el2 is the second column and so on.
   # In oder to prevent data overlap during multiple read/ write opperations, the data should be deleted once used.
   def __init__(self, el1=None, el2=None, el3=None, el4=None, el5=None, el6=None, el7=None, el8=None, el9=None, el10=None, el11=None, el12=None, el13=None, el14=None, el15=None,
                el16=None,el17=None,el18=None,el19=None,el20=None,el21=None,el22=None,el23=None,el24=None,el25=None,el26=None,el27=None,el28=None,el29=None,el30=None,el31=None,el32=None,el33=None,
                el34=None,el35=None,el36=None,el37=None,el38=None,el39=None,el40=None,el41=None,el42=None,el43=None,el44=None,el45=None,el46=None,el47=None,el48=None,el49=None,el50=None):
      self.el1 = el1
      self.el2 = el2
      self.el3 = el3
      self.el4 = el4
      self.el5 = el5
      self.el6 = el6
      self.el7 = el7
      self.el8 = el8
      self.el9 = el9
      self.el10 = el10
      self.el11 = el11
      self.el12 = el12
      self.el13 = el13
      self.el14 = el14
      self.el15 = el15
      self.el16 = el16
      self.el17 = el17
      self.el18 = el18
      self.el19 = el19
      self.el20 = el20
      self.el21 = el21
      self.el22 = el22
      self.el23 = el23
      self.el24 = el24
      self.el25 = el25
      self.el26 = el26
      self.el27 = el27
      self.el28 = el28
      self.el29 = el29
      self.el30 = el30
      self.el31 = el31
      self.el32 = el32
      self.el33 = el33

      self.el34 = el34
      self.el35 = el35
      self.el36 = el36
      self.el37 = el37
      self.el38 = el38
      self.el39 = el39
      self.el40 = el40
      self.el41 = el41
      self.el42 = el42
      self.el43 = el43

      self.el44 = el44
      self.el45 = el45
      self.el46 = el46
      self.el47 = el47

      self.el48 = el48
      self.el49 = el49
      self.el50 = el50
      if self.el1 == None:
         pass
      else:
         Excel.lo_Inst_lst.append(self)


   # Adds an element to the class
   @ classmethod
   def Add_Data(cls, el1, el2=None, el3=None, el4=None, el5=None, el6=None, el7=None,el8=None,el9=None,el10=None,el11=None,el12=None,el13=None,el14=None,el15=None,el16=None, el17=None,el18=None,el19=None,el20=None,el21=None,el22=None,el23=None,el24=None,el25=None,el26=None,el27=None,el28=None,el29=None,el30=None,el31=None,el32=None,
                el33=None,el34=None,el35=None,el36=None,el37=None,el38=None,el39=None,el40=None,el41=None,el42=None,el43=None,el44=None,el45=None,el46=None,el47=None,el48=None,el49=None,el50=None):
      cls(el1, el2, el3, el4, el5, el6, el7, el8, el9, el10, el11, el12, el13, el14, el15, el16, el17, el18, el19, el20, el21, el22, el23, el24, el25, el26, el27, el28, el29, el30, el31, el32,
          el33, el34, el35, el36, el37, el38, el39, el40, el41, el42, el43, el44, el45, el46, el47, el48, el49, el50)


   # Adds an element to the class
   @ classmethod
   def Add_Data_List(cls, llstr_data):
      if len(llstr_data[0]) == 0 or len(llstr_data) == 0:
         print ('Empty list')
      elif len(llstr_data[0]) > 50:
         print ('List to big (maximum 50 columns supported)')

      else:
         if len(llstr_data[0]) == 1:
            for row in llstr_data:
                  cls(row[0])

         elif len(llstr_data[0]) == 2:
            for row in llstr_data:
               cls(row[0],row[1])

         elif len(llstr_data[0]) == 3:
            for row in llstr_data:
               cls(row[0],row[1],row[2])

         elif len(llstr_data[0]) == 4:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3])

         elif len(llstr_data[0]) == 5:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4])

         elif len(llstr_data[0]) == 6:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5])

         elif len(llstr_data[0]) == 7:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6])

         elif len(llstr_data[0]) == 8:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6] ,row[7])

         elif len(llstr_data[0]) == 9:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8])

         elif len(llstr_data[0]) == 10:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9])
         elif len(llstr_data[0]) == 11:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10])

         elif len(llstr_data[0]) == 12:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11])

         elif len(llstr_data[0]) == 13:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12])

         elif len(llstr_data[0]) == 14:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13])

         elif len(llstr_data[0]) == 15:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14])

         elif len(llstr_data[0]) == 16:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15])

         elif len(llstr_data[0]) == 17:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16])

         elif len(llstr_data[0]) == 18:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17])

         elif len(llstr_data[0]) == 19:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18])

         elif len(llstr_data[0]) == 20:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19])

         elif len(llstr_data[0]) == 21:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20])

         elif len(llstr_data[0]) == 22:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21])

         elif len(llstr_data[0]) == 23:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22])

         elif len(llstr_data[0]) == 24:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23])

         elif len(llstr_data[0]) == 25:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24])

         elif len(llstr_data[0]) == 26:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25])

         elif len(llstr_data[0]) == 27:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26])

         elif len(llstr_data[0]) == 28:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27])

         elif len(llstr_data[0]) == 29:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27],row[28])

         elif len(llstr_data[0]) == 30:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27],row[28],row[29])

         elif len(llstr_data[0]) == 31:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27],row[28],row[29],row[30])

         elif len(llstr_data[0]) == 32:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27],row[28],row[29],row[30],row[31])

         elif len(llstr_data[0]) == 33:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27],row[28],row[29],row[30],row[31],row[32])

         elif len(llstr_data[0]) == 34:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27],row[28],row[29],row[30],row[31],row[32],row[33])

         elif len(llstr_data[0]) == 35:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27],row[28],row[29],row[30],row[31],row[32],row[33],row[34])

         elif len(llstr_data[0]) == 36:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27],row[28],row[29],row[30],row[31],row[32],row[33],row[34],row[35])

         elif len(llstr_data[0]) == 37:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27],row[28],row[29],row[30],row[31],row[32],row[33],row[34],row[35],row[36])

         elif len(llstr_data[0]) == 38:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27],row[28],row[29],row[30],row[31],row[32],row[33],row[34],row[35],row[36],row[37])

         elif len(llstr_data[0]) == 39:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27],row[28],row[29],row[30],row[31],row[32],row[33],row[34],row[35],row[36],row[37],row[38])

         elif len(llstr_data[0]) == 40:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27],row[28],row[29],row[30],row[31],row[32],row[33],row[34],row[35],row[36],row[37],row[38],row[39])

         elif len(llstr_data[0]) == 41:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27],row[28],row[29],row[30],row[31],row[32],row[33],row[34],row[35],row[36],row[37],row[38],row[39],row[40])

         elif len(llstr_data[0]) == 42:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27],row[28],row[29],row[30],row[31],row[32],row[33],row[34],row[35],row[36],row[37],row[38],row[39],row[40],row[41])

         elif len(llstr_data[0]) == 43:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27],row[28],row[29],row[30],row[31],row[32],row[33],row[34],row[35],row[36],row[37],row[38],row[39],row[40],row[41],row[42])

         elif len(llstr_data[0]) == 44:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27],row[28],row[29],row[30],row[31],row[32],row[33],row[34],row[35],row[36],row[37],row[38],row[39],row[40],row[41],row[42],row[43])

         elif len(llstr_data[0]) == 45:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27],row[28],row[29],row[30],row[31],row[32],row[33],row[34],row[35],row[36],row[37],row[38],row[39],row[40],row[41],row[42],row[43],row[44])

         elif len(llstr_data[0]) == 46:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27],row[28],row[29],row[30],row[31],row[32],row[33],row[34],row[35],row[36],row[37],row[38],row[39],row[40],row[41],row[42],row[43],row[44],row[45])

         elif len(llstr_data[0]) == 47:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27],row[28],row[29],row[30],row[31],row[32],row[33],row[34],row[35],row[36],row[37],row[38],row[39],row[40],row[41],row[42],row[43],row[44],row[45],row[46])

         elif len(llstr_data[0]) == 48:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27],row[28],row[29],row[30],row[31],row[32],row[33],row[34],row[35],row[36],row[37],row[38],row[39],row[40],row[41],row[42],row[43],row[44],row[45],row[46],row[47])

         elif len(llstr_data[0]) == 49:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27],row[28],row[29],row[30],row[31],row[32],row[33],row[34],row[35],row[36],row[37],row[38],row[39],row[40],row[41],row[42],row[43],row[44],row[45],row[46],row[47], row[48])

         elif len(llstr_data[0]) == 50:
            for row in llstr_data:
               cls(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27],row[28],row[29],row[30],row[31],row[32],row[33],row[34],row[35],row[36],row[37],row[38],row[39],row[40],row[41],row[42],row[43],row[44],row[45],row[46],row[47],row[48],row[49])


   # Removes a class intance based on the given "cls_intance_index" parameter
   @classmethod
   def Remove_Data(cls, cls_intance_index):
      cls.lo_Inst_lst.pop(cls_intance_index)


   # Removes all instances of the class
   @classmethod
   def Remove_All_Data(cls):
      for el in cls.lo_Inst_lst:
         del el
      cls.lo_Inst_lst.clear()


   @classmethod
   # Writes all class intences to the given excel file and sheet.
   # filter_el and filter_val will filter the class data before writing. Only maching data will be written to the excel.
   # filter_el -> number of class parameter: if 1 -> el1, if 2 -> el2 and so on.
   # filter_val -> value to be filterd by(if u pass directly an ID)
   def Write_Sheet(cls, path, sheet, filter_el=None, filter_val=None):
      """
            A method that loads received data into excel file.
      """
      # Verify if the file exists.
      if os.path.exists(path):
         try:
            wb = load_workbook(path)
            if sheet in wb.sheetnames:
               wb.active = wb[sheet]
            else:
               wb.create_sheet(sheet)
               wb.active = wb[sheet]
            if filter_el == None and filter_val == None:
               ws = wb.active
               # Delete everything from active sheet before writing to it.
               while ws.max_row > 1:
                  ws.delete_rows(2)
               # Create the headers.
               ws["A1"] = "Time_Slice(ms)"
               ws["B1"] = "Task1"
               ws["C1"] = "Task2"
               ws["D1"] = "Task3"
               ws["E1"] = "Task4"
               ws["F1"] = "Task5"
               ws["G1"] = "Task6"
               ws["H1"] = "Task7"
               ws["I1"] = "Task8"
               ws["J1"] = "Task9"
               ws["K1"] = "Task10"
               ws["L1"] = "Task11"
               ws["M1"] = "Task12"
               ws["N1"] = "Task13"
               ws["O1"] = "Total_Exec_Time(ms)"
               ws["A1"].font = Font(bold=True)
               ws["B1"].font = Font(bold=True)
               ws["C1"].font = Font(bold=True)
               ws["D1"].font = Font(bold=True)
               ws["E1"].font = Font(bold=True)
               ws["F1"].font = Font(bold=True)
               ws["G1"].font = Font(bold=True)
               ws["H1"].font = Font(bold=True)
               ws["I1"].font = Font(bold=True)
               ws["J1"].font = Font(bold=True)
               ws["K1"].font = Font(bold=True)
               ws["L1"].font = Font(bold=True)
               ws["M1"].font = Font(bold=True)
               ws["N1"].font = Font(bold=True)
               ws["O1"].font = Font(bold=True)


               row = 1
               col = 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el1)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el2)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el3)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el4)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el5)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el6)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el7)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el8)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el9)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el10)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el11)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el12)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el13)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el14)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el15)
                  row += 1
               row = 1
               col += 1
               wb.save(path)
               wb.close()
            elif filter_el != None and filter_val == None:
               ws = wb.active
               # Delete everything from active sheet before writing to it.
               while ws.max_row > 1:
                  ws.delete_rows(2)
               # Create the headers.
               ws["A1"] = "Time_Slice(ms)"
               ws["B1"] = "Task1"
               ws["C1"] = "Task2"
               ws["D1"] = "Task3"
               ws["E1"] = "Task4"
               ws["F1"] = "Task5"
               ws["G1"] = "Task6"
               ws["H1"] = "Task7"
               ws["I1"] = "Task8"
               ws["J1"] = "Task9"
               ws["K1"] = "Task10"
               ws["L1"] = "Task11"
               ws["M1"] = "Task12"
               ws["N1"] = "Task13"
               ws["O1"] = "Total_Exec_Time(ms)"
               ws["A1"].font = Font(bold=True)
               ws["B1"].font = Font(bold=True)
               ws["C1"].font = Font(bold=True)
               ws["D1"].font = Font(bold=True)
               ws["E1"].font = Font(bold=True)
               ws["F1"].font = Font(bold=True)
               ws["G1"].font = Font(bold=True)
               ws["H1"].font = Font(bold=True)
               ws["I1"].font = Font(bold=True)
               ws["J1"].font = Font(bold=True)
               ws["K1"].font = Font(bold=True)
               ws["L1"].font = Font(bold=True)
               ws["M1"].font = Font(bold=True)
               ws["N1"].font = Font(bold=True)
               ws["O1"].font = Font(bold=True)
               # Save and close
               wb.save(path)
               wb.close()
            elif filter_el == None and filter_val != None:
               ws = wb.active
               # Delete everything from active sheet before writing to it.
               while ws.max_row > 1:
                  ws.delete_rows(2)
               # Create the headers.
               ws["A1"] = "Time_Slice(ms)"
               ws["B1"] = "Task1"
               ws["C1"] = "Task2"
               ws["D1"] = "Task3"
               ws["E1"] = "Task4"
               ws["F1"] = "Task5"
               ws["G1"] = "Task6"
               ws["H1"] = "Task7"
               ws["I1"] = "Task8"
               ws["J1"] = "Task9"
               ws["K1"] = "Task10"
               ws["L1"] = "Task11"
               ws["M1"] = "Task12"
               ws["N1"] = "Task13"
               ws["O1"] = "Total_Exec_Time(ms)"
               ws["A1"].font = Font(bold=True)
               ws["B1"].font = Font(bold=True)
               ws["C1"].font = Font(bold=True)
               ws["D1"].font = Font(bold=True)
               ws["E1"].font = Font(bold=True)
               ws["F1"].font = Font(bold=True)
               ws["G1"].font = Font(bold=True)
               ws["H1"].font = Font(bold=True)
               ws["I1"].font = Font(bold=True)
               ws["J1"].font = Font(bold=True)
               ws["K1"].font = Font(bold=True)
               ws["L1"].font = Font(bold=True)
               ws["M1"].font = Font(bold=True)
               ws["N1"].font = Font(bold=True)
               ws["O1"].font = Font(bold=True)
               for i in range(len(cls.lo_Inst_lst)):
                  if str(cls.lo_Inst_lst[i].el1) == str(filter_val):
                     ws["A2"] = cls.lo_Inst_lst[i].el1
                     ws["B2"] = cls.lo_Inst_lst[i].el2
                     ws["C2"] = cls.lo_Inst_lst[i].el3
                     ws["D2"] = cls.lo_Inst_lst[i].el4
                     ws["E2"] = cls.lo_Inst_lst[i].el5
                     ws["F2"] = cls.lo_Inst_lst[i].el6
                     ws["G2"] = cls.lo_Inst_lst[i].el7
                     ws["H2"] = cls.lo_Inst_lst[i].el8
                     ws["I2"] = cls.lo_Inst_lst[i].el9
                     ws["J2"] = cls.lo_Inst_lst[i].el10
                     ws["K2"] = cls.lo_Inst_lst[i].el11
                     ws["L2"] = cls.lo_Inst_lst[i].el12
                     ws["M2"] = cls.lo_Inst_lst[i].el13
                     ws["N2"] = cls.lo_Inst_lst[i].el14
                     ws["O2"] = cls.lo_Inst_lst[i].el15

               # Save and close
               wb.save(path)
               wb.close()
            elif filter_el != None and filter_val != None:
               ws = wb.active
               # Delete everything from active sheet before writing to it.
               while ws.max_row > 1:
                  ws.delete_rows(2)
               # Create the headers.
               ws["A1"] = "Time_Slice(ms)"
               ws["B1"] = "Task1"
               ws["C1"] = "Task2"
               ws["D1"] = "Task3"
               ws["E1"] = "Task4"
               ws["F1"] = "Task5"
               ws["G1"] = "Task6"
               ws["H1"] = "Task7"
               ws["I1"] = "Task8"
               ws["J1"] = "Task9"
               ws["K1"] = "Task10"
               ws["L1"] = "Task11"
               ws["M1"] = "Task12"
               ws["N1"] = "Task13"
               ws["O1"] = "Total_Exec_Time(ms)"
               ws["A1"].font = Font(bold=True)
               ws["B1"].font = Font(bold=True)
               ws["C1"].font = Font(bold=True)
               ws["D1"].font = Font(bold=True)
               ws["E1"].font = Font(bold=True)
               ws["F1"].font = Font(bold=True)
               ws["G1"].font = Font(bold=True)
               ws["H1"].font = Font(bold=True)
               ws["I1"].font = Font(bold=True)
               ws["J1"].font = Font(bold=True)
               ws["K1"].font = Font(bold=True)
               ws["L1"].font = Font(bold=True)
               ws["M1"].font = Font(bold=True)
               ws["N1"].font = Font(bold=True)
               ws["O1"].font = Font(bold=True)
               row = 1
               col = 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el1)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el2)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el3)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el4)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el5)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el6)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el7)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el8)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el9)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el10)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el11)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el12)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el13)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el14)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el15)
               col = 1
               row += 1
               for i in range(len(cls.lo_Inst_lst)):
                  if str(cls.lo_Inst_lst[i].el1) == str(filter_val):
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el1)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el2)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el3)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el4)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el5)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el6)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el7)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el8)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el9)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el10)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el11)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el12)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el13)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el14)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el15)
                     col = 1
                     row = 1
               # Save and close
               wb.save(path)
               wb.close()
         except Exception as e:
               return e
      # If file doesn't exist, create one.
      else:
         wb = Workbook()
         wb.remove_sheet(wb.active)
         wb.active = wb.create_sheet(sheet)
         ws = wb.active
         try:
            # Create the headers.
            ws["A1"] = "Time_Slice(ms)"
            ws["B1"] = "Task1"
            ws["C1"] = "Task2"
            ws["D1"] = "Task3"
            ws["E1"] = "Task4"
            ws["F1"] = "Task5"
            ws["G1"] = "Task6"
            ws["H1"] = "Task7"
            ws["I1"] = "Task8"
            ws["J1"] = "Task9"
            ws["K1"] = "Task10"
            ws["L1"] = "Task11"
            ws["M1"] = "Task12"
            ws["N1"] = "Task13"
            ws["O1"] = "Total_Exec_Time(ms)"
            ws["A1"].font = Font(bold=True)
            ws["B1"].font = Font(bold=True)
            ws["C1"].font = Font(bold=True)
            ws["D1"].font = Font(bold=True)
            ws["E1"].font = Font(bold=True)
            ws["F1"].font = Font(bold=True)
            ws["G1"].font = Font(bold=True)
            ws["H1"].font = Font(bold=True)
            ws["I1"].font = Font(bold=True)
            ws["J1"].font = Font(bold=True)
            ws["K1"].font = Font(bold=True)
            ws["L1"].font = Font(bold=True)
            ws["M1"].font = Font(bold=True)
            ws["N1"].font = Font(bold=True)
            ws["O1"].font = Font(bold=True)
            row = 1
            col = 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el1)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el2)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el3)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el4)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el5)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el6)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el7)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el8)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el9)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el10)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el11)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el12)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el13)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el14)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el15)
               row += 1
            row = 1
            col += 1
            # Save and close
            wb.save(path)
            wb.close()
         except Exception as e:
            return e

   @classmethod
   def Write_Sheet_Including_Headders(cls, path, sheet, filter_el=None, filter_val=None):
      """
            A method that loads received data into excel file.
      """
      # Verify if the file exists.
      if os.path.exists(path):
         try:
            wb = load_workbook(path)
            if sheet in wb.sheetnames:
               wb.active = wb[sheet]
            else:
               wb.create_sheet(sheet)
               wb.active = wb[sheet]
            if filter_el == None and filter_val == None:
               ws = wb.active
               # Delete everything from active sheet before writing to it.
               while ws.max_row > 1:
                  ws.delete_rows(2)
               # Create the headers.
               row = 1
               col = 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el1)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el2)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el3)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el4)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el5)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el6)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el7)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el8)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el9)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el10)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el11)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el12)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el13)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el14)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el15)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el16)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el17)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el18)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el19)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el20)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el21)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el22)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el23)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el24)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el25)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el26)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el27)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el28)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el29)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el30)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el31)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el32)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el33)
                  row += 1
               row = 1
               col += 1


               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el34)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el35)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el36)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el37)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el38)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el39)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el40)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el41)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el42)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el43)
                  row += 1
               row = 1
               col += 1

               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el44)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el45)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el46)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el47)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el48)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el49)
                  row += 1
               row = 1
               col += 1
               for obj in cls.lo_Inst_lst:
                  ws.cell(row+1, col, obj.el50)
                  row += 1
               row = 1
               col += 1

               wb.save(path)
               wb.close()
            elif filter_el != None and filter_val == None:
               ws = wb.active
               # Delete everything from active sheet before writing to it.
               while ws.max_row > 1:
                  ws.delete_rows(2)
               # Create the headers.
               # Save and close
               wb.save(path)
               wb.close()
            elif filter_el == None and filter_val != None:
               ws = wb.active
               # Delete everything from active sheet before writing to it.
               while ws.max_row > 1:
                  ws.delete_rows(2)
               # Create the headers.
               for i in range(len(cls.lo_Inst_lst)):
                  if str(cls.lo_Inst_lst[i].el1) == str(filter_val):
                     ws["A2"] = cls.lo_Inst_lst[i].el1
                     ws["B2"] = cls.lo_Inst_lst[i].el2
                     ws["C2"] = cls.lo_Inst_lst[i].el3
                     ws["D2"] = cls.lo_Inst_lst[i].el4
                     ws["E2"] = cls.lo_Inst_lst[i].el5
                     ws["F2"] = cls.lo_Inst_lst[i].el6
                     ws["G2"] = cls.lo_Inst_lst[i].el7
                     ws["H2"] = cls.lo_Inst_lst[i].el8
                     ws["I2"] = cls.lo_Inst_lst[i].el9
                     ws["J2"] = cls.lo_Inst_lst[i].el10
                     ws["K2"] = cls.lo_Inst_lst[i].el11
                     ws["L2"] = cls.lo_Inst_lst[i].el12
                     ws["M2"] = cls.lo_Inst_lst[i].el13
                     ws["N2"] = cls.lo_Inst_lst[i].el14
                     ws["O2"] = cls.lo_Inst_lst[i].el15
                     ws["P2"] = cls.lo_Inst_lst[i].el16
                     ws["Q2"] = cls.lo_Inst_lst[i].el17
                     ws["R2"] = cls.lo_Inst_lst[i].el18
                     ws["S2"] = cls.lo_Inst_lst[i].el19
                     ws["T2"] = cls.lo_Inst_lst[i].el20
                     ws["U2"] = cls.lo_Inst_lst[i].el21
                     ws["V2"] = cls.lo_Inst_lst[i].el22
                     ws["W2"] = cls.lo_Inst_lst[i].el23
                     ws["X2"] = cls.lo_Inst_lst[i].el24
                     ws["Y2"] = cls.lo_Inst_lst[i].el25
                     ws["Z2"] = cls.lo_Inst_lst[i].el26
                     ws["AA2"] = cls.lo_Inst_lst[i].el27
                     ws["AB2"] = cls.lo_Inst_lst[i].el28
                     ws["AC2"] = cls.lo_Inst_lst[i].el29
                     ws["AD2"] = cls.lo_Inst_lst[i].el30
                     ws["AE2"] = cls.lo_Inst_lst[i].el31
                     ws["AF2"] = cls.lo_Inst_lst[i].el32
                     ws["AG2"] = cls.lo_Inst_lst[i].el33

                     ws["AH2"] = cls.lo_Inst_lst[i].el34
                     ws["AI2"] = cls.lo_Inst_lst[i].el35
                     ws["AJ2"] = cls.lo_Inst_lst[i].el36
                     ws["AK2"] = cls.lo_Inst_lst[i].el37
                     ws["AL2"] = cls.lo_Inst_lst[i].el38
                     ws["AM2"] = cls.lo_Inst_lst[i].el39
                     ws["AN2"] = cls.lo_Inst_lst[i].el40
                     ws["AO2"] = cls.lo_Inst_lst[i].el41
                     ws["AP2"] = cls.lo_Inst_lst[i].el42
                     ws["AQ2"] = cls.lo_Inst_lst[i].el43


                     ws["AR2"] = cls.lo_Inst_lst[i].el44
                     ws["AS2"] = cls.lo_Inst_lst[i].el45
                     ws["AT2"] = cls.lo_Inst_lst[i].el46
                     ws["AU2"] = cls.lo_Inst_lst[i].el47
                     ws["AV2"] = cls.lo_Inst_lst[i].el48
                     ws["AW2"] = cls.lo_Inst_lst[i].el49
                     ws["AX2"] = cls.lo_Inst_lst[i].el50

               # Save and close
               wb.save(path)
               wb.close()
            elif filter_el != None and filter_val != None:
               ws = wb.active
               # Delete everything from active sheet before writing to it.
               while ws.max_row > 1:
                  ws.delete_rows(2)
               # Create the headers.

               row = 1
               col = 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el1)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el2)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el3)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el4)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el5)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el6)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el7)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el8)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el9)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el10)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el11)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el12)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el13)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el14)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el15)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el16)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el17)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el18)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el19)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el20)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el21)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el22)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el23)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el24)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el25)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el26)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el27)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el28)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el29)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el30)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el31)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el32)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el33)

               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el34)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el35)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el36)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el37)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el38)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el39)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el40)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el41)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el42)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el43)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el44)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el45)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el46)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el47)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el48)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el49)
               col += 1
               ws.cell(row + 1, col, cls.lo_Inst_lst[filter_el].el50)

               col = 1
               row += 1
               for i in range(len(cls.lo_Inst_lst)):
                  if str(cls.lo_Inst_lst[i].el1) == str(filter_val):
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el1)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el2)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el3)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el4)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el5)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el6)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el7)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el8)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el9)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el10)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el11)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el12)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el13)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el14)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el15)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el16)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el17)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el18)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el19)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el20)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el21)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el22)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el23)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el24)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el25)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el26)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el27)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el28)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el29)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el30)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el31)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el32)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el33)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el34)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el35)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el36)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el37)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el38)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el39)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el40)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el41)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el42)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el43)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el44)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el45)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el46)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el47)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el48)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el49)
                     col += 1
                     ws.cell(row + 1, col, cls.lo_Inst_lst[i].el50)
                     col = 1
                     row = 1
               # Save and close
               wb.save(path)
               wb.close()
         except Exception as e:
            return e
      # If file doesn't exist, create one.
      else:
         wb = Workbook()
         wb.remove_sheet(wb.active)
         wb.active = wb.create_sheet(sheet)
         ws = wb.active
         try:
            # Create the headers.

            row = 1
            col = 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el1)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el2)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el3)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el4)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el5)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el6)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el7)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el8)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el9)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el10)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el11)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el12)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el13)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el14)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row + 1, col, obj.el15)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el16)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el17)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el18)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el19)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el20)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el21)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el22)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el23)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el24)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el25)
               row += 1
            row = 1
            col += 1

            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el26)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el27)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el28)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el29)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el30)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el31)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el32)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el33)
               row += 1
            row = 1
            col += 1

            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el34)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el35)
               row += 1
            row = 1
            col += 1

            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el36)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el37)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el38)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el39)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el40)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el41)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el42)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el43)
               row += 1
            row = 1
            col += 1

            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el44)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el45)
               row += 1
            row = 1
            col += 1

            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el46)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el47)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el48)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el49)
               row += 1
            row = 1
            col += 1
            for obj in cls.lo_Inst_lst:
               ws.cell(row+1, col, obj.el50)
               row += 1
            row = 1
            col += 1


            # Save and close
            wb.save(path)
            wb.close()
         except Exception as e:
            return e


   @classmethod
   def Import_Sheet(cls, path, sheet):
      """
      Get everything from the excel sheet and add each line to the class.
      Each column element shall be mapped as class's elements as follows:
      - column A = el1, column B = el2, etc.
      - row 0 = instance 0, row 1 = instance 1, etc.
      """
      if os.path.exists(path):
         wb = load_workbook(path)
         worksheets = wb.worksheets
         if wb[sheet] in worksheets:
            ws = wb.active
            wb.active = wb[sheet]
            for col in ws.iter_rows(min_row=2, min_col=ws.min_column, max_col=ws.max_column):
               if ws.max_column == 1:
                  cls(col[0].value)
               if ws.max_column == 2:
                  cls(col[0].value, col[1].value)
               if ws.max_column == 3:
                  cls(col[0].value, col[1].value, col[2].value)
               if ws.max_column == 4:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value)
               if ws.max_column == 5:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value)
               if ws.max_column == 6:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value)
               if ws.max_column == 7:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value)
               if ws.max_column == 8:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value)
               if ws.max_column == 9:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value)
               if ws.max_column == 10:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value)
               if ws.max_column == 11:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value)
               if ws.max_column == 12:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value)
               if ws.max_column == 13:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value)
               if ws.max_column == 14:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value)
               if ws.max_column == 15:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value)
               if ws.max_column == 16:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value)
               if ws.max_column == 17:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value)
               if ws.max_column == 18:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value)
               if ws.max_column == 19:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value)
               if ws.max_column == 20:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value)
               if ws.max_column == 21:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value)
               if ws.max_column == 22:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value)
               if ws.max_column == 23:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value)
               if ws.max_column == 24:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value)
               if ws.max_column == 25:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value)
               if ws.max_column == 26:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value)
               if ws.max_column == 27:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value)
               if ws.max_column == 28:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value, col[27].value)
               if ws.max_column == 29:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value, col[27].value, col[28].value)
               if ws.max_column == 30:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value, col[27].value, col[28].value, col[29].value)
               if ws.max_column == 31:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value, col[27].value, col[28].value, col[29].value, col[30].value)
               if ws.max_column == 32:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value, col[27].value, col[28].value, col[29].value, col[30].value, col[31].value)
               if ws.max_column == 33:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value, col[27].value, col[28].value, col[29].value, col[30].value, col[31].value, col[32].value)
               if ws.max_column == 34:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value, col[27].value, col[28].value, col[29].value, col[30].value, col[31].value, col[32].value, col[33].value)
               if ws.max_column == 35:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value, col[27].value, col[28].value, col[29].value, col[30].value, col[31].value, col[32].value, col[33].value, col[34].value)
               if ws.max_column == 36:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value, col[27].value, col[28].value, col[29].value, col[30].value, col[31].value, col[32].value, col[33].value, col[34].value, col[35].value)
               if ws.max_column == 37:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value, col[27].value, col[28].value, col[29].value, col[30].value, col[31].value, col[32].value, col[33].value, col[34].value, col[35].value, col[36].value)
               if ws.max_column == 38:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value, col[27].value, col[28].value, col[29].value, col[30].value, col[31].value, col[32].value, col[33].value, col[34].value, col[35].value, col[36].value, col[37].value)
               if ws.max_column == 39:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value, col[27].value, col[28].value, col[29].value, col[30].value, col[31].value, col[32].value, col[33].value, col[34].value, col[35].value, col[36].value, col[37].value, col[38].value)
               if ws.max_column == 40:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value, col[27].value, col[28].value, col[29].value, col[30].value, col[31].value, col[32].value, col[33].value, col[34].value, col[35].value, col[36].value, col[37].value, col[38].value, col[39].value)
               if ws.max_column == 41:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value, col[27].value, col[28].value, col[29].value, col[30].value, col[31].value, col[32].value, col[33].value, col[34].value, col[35].value, col[36].value, col[37].value, col[38].value, col[39].value, col[40].value)
               if ws.max_column == 42:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value, col[27].value, col[28].value, col[29].value, col[30].value, col[31].value, col[32].value, col[33].value, col[34].value, col[35].value, col[36].value, col[37].value, col[38].value, col[39].value, col[40].value, col[41].value)
               if ws.max_column == 43:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value, col[27].value, col[28].value, col[29].value, col[30].value, col[31].value, col[32].value, col[33].value, col[34].value, col[35].value, col[36].value, col[37].value, col[38].value, col[39].value, col[40].value, col[41].value, col[42].value)
               if ws.max_column == 44:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value, col[27].value, col[28].value, col[29].value, col[30].value, col[31].value, col[32].value, col[33].value, col[34].value, col[35].value, col[36].value, col[37].value, col[38].value, col[39].value, col[40].value, col[41].value, col[42].value, col[43].value)
               if ws.max_column == 45:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value, col[27].value, col[28].value, col[29].value, col[30].value, col[31].value, col[32].value, col[33].value, col[34].value, col[35].value, col[36].value, col[37].value, col[38].value, col[39].value, col[40].value, col[41].value, col[42].value, col[43].value, col[44].value)
               if ws.max_column == 46:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value, col[27].value, col[28].value, col[29].value, col[30].value, col[31].value, col[32].value, col[33].value, col[34].value, col[35].value, col[36].value, col[37].value, col[38].value, col[39].value, col[40].value, col[41].value, col[42].value, col[43].value, col[44].value, col[45].value)
               if ws.max_column == 47:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value, col[27].value, col[28].value, col[29].value, col[30].value, col[31].value, col[32].value, col[33].value, col[34].value, col[35].value, col[36].value, col[37].value, col[38].value, col[39].value, col[40].value, col[41].value, col[42].value, col[43].value, col[44].value, col[45].value, col[46].value)
               if ws.max_column == 48:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value, col[27].value, col[28].value, col[29].value, col[30].value, col[31].value, col[32].value, col[33].value, col[34].value, col[35].value, col[36].value, col[37].value, col[38].value, col[39].value, col[40].value, col[41].value, col[42].value, col[43].value, col[44].value, col[45].value, col[46].value, col[47].value)
               if ws.max_column == 49:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value, col[27].value, col[28].value, col[29].value, col[30].value, col[31].value, col[32].value, col[33].value, col[34].value, col[35].value, col[36].value, col[37].value, col[38].value, col[39].value, col[40].value, col[41].value, col[42].value, col[43].value, col[44].value, col[45].value, col[46].value, col[47].value, col[48].value)
               if ws.max_column >= 50:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value, col[7].value, col[8].value, col[9].value, col[10].value, col[11].value, col[12].value, col[13].value, col[14].value, col[15].value, col[16].value, col[17].value, col[18].value, col[19].value, col[20].value, col[21].value, col[22].value, col[23].value, col[24].value, col[25].value, col[26].value, col[27].value, col[28].value, col[29].value, col[30].value, col[31].value, col[32].value, col[33].value, col[34].value, col[35].value, col[36].value, col[37].value, col[38].value, col[39].value, col[40].value, col[41].value, col[42].value, col[43].value, col[44].value, col[45].value, col[46].value, col[47].value, col[48].value, col[49].value)
            return "No_Error"
         else:
            return "Worksheet not found."
      else:
         return "There's no file at the given path."


   @classmethod
   def Get_Excel_Data(cls):
      return Excel.lo_Inst_lst

   @classmethod
   def Create_Excel_File(cls,str_path,str_name,lstr_sheets=None):
      wb = Workbook()
      ws =  wb.active
      ws.title = "Sheet1"
      if lstr_sheets != None:
         for el in lstr_sheets:
            wb.active = wb.create_sheet(el)
      wb.save(filename = os.path.join(str_path,str_name))
      return os.path.join(str_path,str_name)